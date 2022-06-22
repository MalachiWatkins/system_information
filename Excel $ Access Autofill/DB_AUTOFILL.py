import ctypes
from ctypes import wintypes
import time
import json
import pyperclip as pc
import keyboard
import platform
#import pylightxl as xl
import yaml
import openpyxl
import datetime
from datetime import date
today = date.today()
from openpyxl import load_workbook
user32 = ctypes.WinDLL('user32', use_last_error=True)
with open("config.yml", "r") as ymlfile:
    cfg = yaml.safe_load(ymlfile)
    json_file_location = cfg['cfg']['json_file_location']
    COA_hotkey = cfg['cfg']['COA_hotkey']
    refurb_hotkey = cfg['cfg']['refurbDB_hotkey']
INPUT_MOUSE = 0
INPUT_KEYBOARD = 1
INPUT_HARDWARE = 2

KEYEVENTF_EXTENDEDKEY = 0x0001
KEYEVENTF_KEYUP       = 0x0002
KEYEVENTF_UNICODE     = 0x0004
KEYEVENTF_SCANCODE    = 0x0008

MAPVK_VK_TO_VSC = 0

# msdn.microsoft.com/en-us/library/dd375731
VK_TAB  = 0x09
VK_MENU = 0x12
VK_P = 0x50
VK_CONTROL = 0x11
VK_V = 0x56
VK_RETURN = 0x0D
# C struct definitions

wintypes.ULONG_PTR = wintypes.WPARAM

class MOUSEINPUT(ctypes.Structure):
    _fields_ = (("dx",          wintypes.LONG),
                ("dy",          wintypes.LONG),
                ("mouseData",   wintypes.DWORD),
                ("dwFlags",     wintypes.DWORD),
                ("time",        wintypes.DWORD),
                ("dwExtraInfo", wintypes.ULONG_PTR))

class KEYBDINPUT(ctypes.Structure):
    _fields_ = (("wVk",         wintypes.WORD),
                ("wScan",       wintypes.WORD),
                ("dwFlags",     wintypes.DWORD),
                ("time",        wintypes.DWORD),
                ("dwExtraInfo", wintypes.ULONG_PTR))

    def __init__(self, *args, **kwds):
        super(KEYBDINPUT, self).__init__(*args, **kwds)
        # some programs use the scan code even if KEYEVENTF_SCANCODE
        # isn't set in dwFflags, so attempt to map the correct code.
        if not self.dwFlags & KEYEVENTF_UNICODE:
            self.wScan = user32.MapVirtualKeyExW(self.wVk,
                                                 MAPVK_VK_TO_VSC, 0)

class HARDWAREINPUT(ctypes.Structure):
    _fields_ = (("uMsg",    wintypes.DWORD),
                ("wParamL", wintypes.WORD),
                ("wParamH", wintypes.WORD))

class INPUT(ctypes.Structure):
    class _INPUT(ctypes.Union):
        _fields_ = (("ki", KEYBDINPUT),
                    ("mi", MOUSEINPUT),
                    ("hi", HARDWAREINPUT))
    _anonymous_ = ("_input",)
    _fields_ = (("type",   wintypes.DWORD),
                ("_input", _INPUT))

LPINPUT = ctypes.POINTER(INPUT)

def _check_count(result, func, args):
    if result == 0:
        raise ctypes.WinError(ctypes.get_last_error())
    return args

user32.SendInput.errcheck = _check_count
user32.SendInput.argtypes = (wintypes.UINT, # nInputs
                             LPINPUT,       # pInputs
                             ctypes.c_int)  # cbSize

# Functions

def PressKey(hexKeyCode):
    x = INPUT(type=INPUT_KEYBOARD,
              ki=KEYBDINPUT(wVk=hexKeyCode))
    user32.SendInput(1, ctypes.byref(x), ctypes.sizeof(x))

def ReleaseKey(hexKeyCode):
    x = INPUT(type=INPUT_KEYBOARD,
              ki=KEYBDINPUT(wVk=hexKeyCode,
                            dwFlags=KEYEVENTF_KEYUP))
    user32.SendInput(1, ctypes.byref(x), ctypes.sizeof(x))

def DB_AUTOFILL():
    json_File = str(json_file_location) + 'parsed.json' # change path if needed
    with open(json_File, 'r') as openfile:
        json_object = json.load(openfile)
    x=0
    while x < 16:
        a = json_object[x]
        list_of_dict_values = list(a.values())
        value = list_of_dict_values[0]
        if x == 15:
            a = pc.copy(value)
            pc.paste()
            print(pc.paste())
            time.sleep(.2)
            PressKey(VK_CONTROL)
            PressKey(VK_V)
            ReleaseKey(VK_CONTROL)
            ReleaseKey(VK_V)
            PressKey(VK_MENU)
            PressKey(VK_P)
            ReleaseKey(VK_MENU)
            ReleaseKey(VK_P)
            PressKey(VK_RETURN)
            ReleaseKey(VK_RETURN)
            time.sleep(0.2)

        else:
            a = pc.copy(value)
            pc.paste()
            print(pc.paste())
            time.sleep(.2)
            PressKey(VK_CONTROL)
            PressKey(VK_V)
            ReleaseKey(VK_CONTROL)
            ReleaseKey(VK_V)
            PressKey(VK_TAB)
            ReleaseKey(VK_TAB)
            time.sleep(0.2)
        x+=1

def COA_AUTOFILL():
    with open('G:\My Drive\ECOM-COMPUTERS\Databases\DONTDELETE.txt', "r") as coa_coll:
        coa_collum = coa_coll.readline()
    wb = load_workbook('G:\My Drive\ECOM-COMPUTERS\Databases\FY21 MAR Citizenship COA report 2022 (April).xlsx')
    sheet = wb.active 
    date = sheet.cell(row = int(coa_collum), column = 5) 
    date.value = str(today)
    y = coa_collum
    x = 0
    json_File = str(json_file_location) + 'parsed.json' # change path if needed
    with open(json_File, 'r') as openfile:
        json_object = json.load(openfile)
    while x < 17:
        a = json_object[x]
        list_of_dict_values = list(a.values())
        value = list_of_dict_values[0]
        if x == 16:
            devType = sheet.cell(row = int(coa_collum), column = 7) 
            devType.value = str(value)
        
        elif x == 0:
            SN = sheet.cell(row = int(coa_collum), column = 8) 
            SN.value = str(value)
        elif x == 1:
            brand = sheet.cell(row = int(coa_collum), column = 9) 
            brand.value = str(value)
        elif x == 2:
            mod = sheet.cell(row = int(coa_collum), column = 10) 
            mod.value = str(value)
        elif x == 3:
            pro = sheet.cell(row = int(coa_collum), column = 11) 
            pro.value = str(value)
        elif x == 11:
            lic = sheet.cell(row = int(coa_collum), column = 14) 
            lic.value = str(value)

        x+=1
    wb.save(filename='G:\My Drive\ECOM-COMPUTERS\Databases\FY21 MAR Citizenship COA report 2022 (April).xlsx')
    # Date:E Device Type:G SN:H Manufac:I model:J CPU:K COA:N
    print(y)
    with open('G:\My Drive\ECOM-COMPUTERS\Databases\DONTDELETE.txt', "w") as coa_coll:
        empty = int(y) + 1
        coa_coll.write(str(empty))
    return
if __name__ == "__main__":
    while True:
        keyboard.add_hotkey(str(refurb_hotkey), lambda: DB_AUTOFILL())
        keyboard.add_hotkey(str(COA_hotkey), lambda: COA_AUTOFILL())
        keyboard.wait()

    #HOTKEYCOMMANDGOESHERE
